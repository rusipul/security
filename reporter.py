from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 색상 팔레트 ──────────────────────────────────────────────
_NAVY        = "1F4E79"
_RED         = "C00000"
_RED_BG      = "FFCCCC"
_RED_CARD    = "FF4444"
_ORANGE_BG   = "FFE5CC"
_ORANGE_CARD = "FF8C00"
_YELLOW_BG   = "FFF2CC"
_YELLOW_CARD = "FFC000"
_GREEN_BG    = "E2EFDA"
_GRAY_BG     = "F5F5F5"
_GRAY2       = "D9D9D9"
_WHITE       = "FFFFFF"
_NAVY_LIGHT  = "D6E4F0"

_F_WHITE   = Font(color=_WHITE,  bold=True,  size=10)
_F_NAVY    = Font(color=_NAVY,   bold=True,  size=10)
_F_TITLE   = Font(color=_NAVY,   bold=True,  size=15)
_F_SECTION = Font(color=_WHITE,  bold=True,  size=10)
_F_BODY    = Font(size=10)
_F_BOLD    = Font(bold=True,     size=10)
_F_SMALL   = Font(size=9,        color="595959")
_F_RED     = Font(color=_RED,    bold=True,  size=10)

_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_LEFT_TOP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

_S_THIN  = Side(style="thin",   color="BFBFBF")
_S_MED   = Side(style="medium", color=_NAVY)
_BORDER  = Border(left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_THIN)
_BORDER_T = Border(left=_S_THIN, right=_S_THIN, top=_S_MED,  bottom=_S_THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _cell(ws, row, col, value="", font=None, fill=None, align=None, border=None):
    c = ws.cell(row, col, value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


def _hdr(ws, row, cols, labels, bg=_NAVY):
    for col, label in zip(cols, labels):
        _cell(ws, row, col, label,
              font=_F_WHITE, fill=_fill(bg), align=_CENTER, border=_BORDER)
    ws.row_dimensions[row].height = 18


def _section_bar(ws, row, c1, c2, text, bg=_NAVY):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    _cell(ws, row, c1, text, font=_F_SECTION, fill=_fill(bg), align=_LEFT)
    ws.row_dimensions[row].height = 20


# ── 경보 카드 (2행짜리 강조 박스) ────────────────────────────
def _alert_card(ws, row, col, label, count, card_color, icon):
    """col ~ col+1 범위 2행 카드."""
    # 라벨 행
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
    c = ws.cell(row, col, f"{icon}  {label}")
    c.font      = Font(color=_WHITE, bold=True, size=9)
    c.fill      = _fill(card_color)
    c.alignment = _CENTER
    c.border    = _BORDER
    ws.row_dimensions[row].height = 20

    # 숫자 행
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    c2 = ws.cell(row+1, col, f"{count:,}건")
    c2.font      = Font(color=_WHITE, bold=True, size=18)
    c2.fill      = _fill(card_color)
    c2.alignment = _CENTER
    c2.border    = _BORDER
    ws.row_dimensions[row+1].height = 32


# ── 상세 시트 ────────────────────────────────────────────────
def _write_df_sheet(wb, sheet_name, df, fill_color=_GRAY_BG):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["데이터 없음"])
        return
    for j, h in enumerate(df.columns, 1):
        _cell(ws, 1, j, h, font=_F_WHITE, fill=_fill(_NAVY), align=_CENTER)
    bg = _fill(fill_color)
    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for j, val in enumerate(row_data, 1):
            c = ws.cell(ri, j, val)
            c.fill = bg; c.alignment = _CENTER; c.border = _BORDER
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 50)


# ── 첫 장: 경영진 보고용 ──────────────────────────────────────
def _write_executive_sheet(ws, period, filename,
                           threat_counts, block_counts,
                           risk_table, ai_summary):

    has_ai = bool(ai_summary) and not ai_summary.startswith("(API 키")

    # 페이지 설정
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9        # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth  = 1
    ws.sheet_properties.pageSetUpPr.fitPage = True
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.6
    ws.page_margins.bottom = 0.6

    # 컬럼 폭: A-J (10열)
    # A  B  [C]  D  E  [F]  G  H  [I]  J
    # 좌테이블 3열 / 스페이서 / 우테이블 3열 / 스페이서 / 사용자테이블용 확장
    widths = [20, 8, 11,   2,   20, 8, 11,   2,   10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # ══ 제목 ══════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, f"정보보안 현황 보고서  |  {period}",
          font=_F_TITLE, fill=_fill(_NAVY_LIGHT), align=_CENTER)
    ws.row_dimensions[r].height = 36
    r += 1

    ws.merge_cells(f"A{r}:J{r}")
    now = datetime.now().strftime("%Y년 %m월 %d일  %H:%M")
    _cell(ws, r, 1, f"분석일시: {now}     원본파일: {filename}",
          font=_F_SMALL, fill=_fill(_GRAY_BG), align=_CENTER)
    ws.row_dimensions[r].height = 15
    r += 1
    ws.row_dimensions[r].height = 6; r += 1   # 여백

    # ══ 섹션 1: 경보 카드 ════════════════════════════════════
    _section_bar(ws, r, 1, 10, "  ⚠  보안 경보 현황  —  즉시 확인 필요 항목")
    r += 1

    cards = [
        ("USB 파일복사 차단",  block_counts.get("USB 파일복사 차단", 0),  _RED_CARD,    "🔴"),
        ("파일전송 차단",       block_counts.get("파일전송 차단", 0),       _RED_CARD,    "🔴"),
        ("설계파일 첨부 탐지",  threat_counts.get("설계파일 첨부", 0),      _ORANGE_CARD, "🟠"),
        ("웹사이트 접속 차단",  block_counts.get("웹사이트 접속 차단", 0),  _YELLOW_CARD, "⚠️"),
    ]
    # 카드 4개: 열 1-2 / 3-4 / 6-7 / 8-9 (5, 10열은 여백)
    card_cols = [1, 3, 6, 8]
    for (label, cnt, color, icon), col in zip(cards, card_cols):
        _alert_card(ws, r, col, label, cnt, color, icon)

    r += 2   # 카드 2행 차지
    ws.row_dimensions[r].height = 6; r += 1   # 여백

    # ══ 섹션 2: 차단현황(좌) + 위협탐지(우) ══════════════════
    # 좌: A-C, 우: E-G
    _section_bar(ws, r, 1, 3,
                 "  ① DLP 차단 현황  (시스템 자동 처리)")
    _section_bar(ws, r, 5, 7,
                 "  ② 추가 위협 탐지  (분석기 검출)")
    r += 1

    _hdr(ws, r, [1, 2, 3], ["차단 유형", "건수", "위험도"])
    _hdr(ws, r, [5, 6, 7], ["탐지 유형", "건수", "판단 근거"])
    r += 1

    block_data = [
        ("USB 파일복사 차단",  block_counts.get("USB 파일복사 차단", 0),  "🔴 고위험", _RED_BG),
        ("파일전송 차단",       block_counts.get("파일전송 차단", 0),       "🔴 고위험", _RED_BG),
        ("웹사이트 접속 차단",  block_counts.get("웹사이트 접속 차단", 0),  "🟠 주의",   _ORANGE_BG),
        ("웹사이트 접속 경고",  block_counts.get("웹사이트 접속 경고", 0),  "⚠️ 경고",  _YELLOW_BG),
    ]
    threat_data = [
        ("USB 시도",       threat_counts.get("USB 시도", 0),
         "USB 기기 연결 시도",               _RED_BG),
        ("설계파일 첨부",   threat_counts.get("설계파일 첨부", 0),
         ".sch/.tar 등 설계파일 첨부 감지",   _RED_BG),
        ("대용량 ZIP 첨부", threat_counts.get("대용량 ZIP 첨부", 0),
         "10MB↑ ZIP — 대량자료 반출 의심",   _ORANGE_BG),
        ("AI 플랫폼 접속",  threat_counts.get("AI 플랫폼 접속", 0),
         "ChatGPT/Gemini 기밀 업로드 우려",  _ORANGE_BG),
        ("외부메일 시도",   threat_counts.get("외부메일 시도", 0),
         "네이버/구글메일 접속 시도",          _RED_BG),
        ("메신저 파일전송", threat_counts.get("메신저 파일전송", 0),
         "WeChat/KakaoTalk 파일 전송",       _YELLOW_BG),
    ]

    for i in range(max(len(block_data), len(threat_data))):
        ws.row_dimensions[r].height = 16
        if i < len(block_data):
            label, cnt, risk, bg = block_data[i]
            _cell(ws, r, 1, label, font=_F_BODY, fill=_fill(bg), align=_LEFT,   border=_BORDER)
            _cell(ws, r, 2, f"{cnt:,}건", font=_F_BOLD, fill=_fill(bg), align=_CENTER, border=_BORDER)
            _cell(ws, r, 3, risk,  font=_F_BODY, fill=_fill(bg), align=_CENTER, border=_BORDER)
        if i < len(threat_data):
            label, cnt, reason, bg = threat_data[i]
            _cell(ws, r, 5, label,  font=_F_BODY, fill=_fill(bg), align=_LEFT,   border=_BORDER)
            _cell(ws, r, 6, f"{cnt:,}건", font=_F_BOLD, fill=_fill(bg), align=_CENTER, border=_BORDER)
            _cell(ws, r, 7, reason, font=_F_SMALL, fill=_fill(bg), align=_LEFT,  border=_BORDER)
        r += 1

    ws.row_dimensions[r].height = 6; r += 1   # 여백

    # ══ 섹션 3: 상위 위험 사용자 ══════════════════════════════
    _section_bar(ws, r, 1, 10, "  ③  상위 위험 사용자 현황")
    r += 1

    if not risk_table.empty:
        headers = list(risk_table.columns)
        for j, h in enumerate(headers, 1):
            _cell(ws, r, j, h, font=_F_WHITE, fill=_fill(_NAVY),
                  align=_CENTER, border=_BORDER)
        ws.row_dimensions[r].height = 18
        r += 1
        for row_data in risk_table.itertuples(index=False):
            ws.row_dimensions[r].height = 16
            score = row_data[-1]
            bg = _RED_BG if score >= 10 else _ORANGE_BG if score >= 5 else _GREEN_BG
            for j, val in enumerate(row_data, 1):
                _cell(ws, r, j, val, font=_F_BODY,
                      fill=_fill(bg), align=_CENTER, border=_BORDER)
            r += 1
    else:
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, "탐지된 위험 사용자 없음", font=_F_BODY,
              fill=_fill(_GRAY_BG), align=_CENTER)
        ws.row_dimensions[r].height = 16
        r += 1

    # ══ 섹션 4: AI 분석 (API 키 있을 때만) ═══════════════════
    if has_ai:
        ws.row_dimensions[r].height = 6; r += 1
        _section_bar(ws, r, 1, 10,
                     "  ④  AI 보안 위협 분석  (Claude Sonnet)")
        r += 1
        ws.merge_cells(f"A{r}:J{r + 6}")
        c = ws.cell(r, 1, ai_summary)
        c.font      = Font(size=9)
        c.alignment = _LEFT_TOP
        c.fill      = _fill(_GRAY_BG)
        ws.row_dimensions[r].height = 110


# ── 메인 ─────────────────────────────────────────────────────
def write_report(
    output_path: str,
    period: str,
    filename: str,
    threat_counts: dict,
    block_counts: dict,
    risk_table: pd.DataFrame,
    ai_summary: str,
    usb_df: pd.DataFrame,
    design_df: pd.DataFrame,
    large_zip_df: pd.DataFrame,
    ai_df: pd.DataFrame,
    email_df: pd.DataFrame,
    messenger_df: pd.DataFrame,
    usb_block_df: pd.DataFrame,
    attach_block_df: pd.DataFrame,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "보안위협 요약"

    _write_executive_sheet(
        ws, period, filename,
        threat_counts, block_counts,
        risk_table, ai_summary,
    )

    _write_df_sheet(wb, "USB시도 상세",      usb_df,          _RED_BG)
    _write_df_sheet(wb, "설계파일 상세",     design_df,       _RED_BG)
    _write_df_sheet(wb, "대용량ZIP 상세",    large_zip_df,    _ORANGE_BG)
    _write_df_sheet(wb, "AI플랫폼 상세",     ai_df,           _ORANGE_BG)
    _write_df_sheet(wb, "외부메일 상세",     email_df,        _RED_BG)
    _write_df_sheet(wb, "메신저 상세",       messenger_df,    _YELLOW_BG)
    _write_df_sheet(wb, "USB복사차단 상세",  usb_block_df,    _RED_BG)
    _write_df_sheet(wb, "파일전송차단 상세", attach_block_df, _ORANGE_BG)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
